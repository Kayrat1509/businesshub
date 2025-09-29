from django.contrib import admin
from django import forms
from django.forms import TextInput, Textarea, NumberInput, Select, CheckboxInput
from django.utils.html import format_html
from import_export.admin import ImportExportModelAdmin
from import_export.formats.base_formats import XLSX

from .models import Product, ProductImage
from .forms import ProductAdminForm
from .resources import ProductResource




@admin.register(Product)
class ProductAdmin(ImportExportModelAdmin):
    # Подключаем ресурс для импорта/экспорта
    resource_class = ProductResource
    form = ProductAdminForm

    # Ограничиваем форматы импорта/экспорта только Excel (.xlsx)
    formats = [XLSX]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Принудительно устанавливаем наш шаблон после инициализации ImportExportModelAdmin
        self.change_list_template = 'admin/products/product/change_list.html'
    list_display = [
        "title",
        "company",
        "category", 
        "get_price_display",
        "get_price_conversions",
        "is_service",
        "in_stock",
        "is_active",
        "created_at"
    ]
    list_filter = [
        "is_service",
        "currency", 
        "in_stock",
        "is_active",
        "category",
        "company",
        "created_at",
    ]
    search_fields = ["title", "description", "sku", "company__name"]
    list_editable = ["in_stock", "is_active"]
    list_per_page = 25

    fieldsets = (
        (
            "Основная информация",
            {"fields": ("company", "title", "sku", "description", "category")},
        ),
        (
            "Цена и наличие",
            {"fields": ("price", "currency", "is_service", "in_stock")},
        ),
        ("Настройки публикации", {"fields": ("is_active",)}),
    )
    
    def get_user_company(self, request):
        """Получить компанию текущего пользователя"""
        try:
            from app.companies.models import Company
            return Company.objects.get(owner=request.user)
        except Company.DoesNotExist:
            return None
    
    def get_queryset(self, request):
        """Фильтрация продуктов по компании пользователя"""
        queryset = super().get_queryset(request)
        
        # Суперпользователи видят все продукты
        if request.user.is_superuser:
            return queryset
        
        # Обычные пользователи видят только продукты своей компании
        user_company = self.get_user_company(request)
        if user_company:
            return queryset.filter(company=user_company)
        
        # Если у пользователя нет компании, показываем пустой список
        return queryset.none()
    
    def save_model(self, request, obj, form, change):
        """Привязка продукта к компании при сохранении"""
        # Суперпользователи могут сохранять с любой компанией
        if request.user.is_superuser:
            super().save_model(request, obj, form, change)
            return
        
        # Для обычных пользователей при создании нового продукта
        # обязательно привязываем к их компании
        if not change:  # Новый продукт
            user_company = self.get_user_company(request)
            if user_company:
                obj.company = user_company
            else:
                # Если у пользователя нет компании, не даем создать продукт
                from django.core.exceptions import PermissionDenied
                raise PermissionDenied("У вас нет компании для создания продуктов")
        
        super().save_model(request, obj, form, change)
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """Ограничение выбора компании и настройка категорий в форме"""
        if db_field.name == "company":
            if request.user.is_superuser:
                # Суперпользователи видят все компании
                pass  # Оставляем queryset без изменений
            else:
                # Обычные пользователи видят только свою компанию
                user_company = self.get_user_company(request)
                if user_company:
                    from app.companies.models import Company
                    kwargs["queryset"] = Company.objects.filter(id=user_company.id)
                else:
                    # Если нет компании, показываем пустой список
                    kwargs["queryset"] = db_field.related_model.objects.none()
        
        elif db_field.name == "category":
            # Настраиваем категории с автоматическим добавлением "Другое"
            self._ensure_other_categories()
        
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    def _ensure_other_categories(self):
        """Автоматически добавляет 'Другое' в каждую группу категорий, если его нет"""
        from app.categories.models import Category
        
        # Получаем все родительские категории (верхний уровень)
        parent_categories = Category.objects.filter(parent__isnull=True, is_active=True)
        
        for parent in parent_categories:
            # Проверяем, есть ли уже "Другое" среди дочерних категорий
            has_other = parent.children.filter(name="Другое", is_active=True).exists()
            
            if not has_other:
                # Создаем "Другое" для этой родительской категории
                Category.objects.get_or_create(
                    name="Другое",
                    parent=parent,
                    defaults={
                        'is_active': True,
                        'slug': f'other-{parent.slug}' if parent.slug else f'other-{parent.id}'
                    }
                )
            
            # Проверяем дочерние категории, которые сами являются родительскими
            child_categories = parent.children.filter(is_active=True)
            for child in child_categories:
                # Если у дочерней категории есть свои дочерние элементы
                if child.children.exists():
                    child_has_other = child.children.filter(name="Другое", is_active=True).exists()
                    
                    if not child_has_other:
                        # Создаем "Другое" для этой дочерней категории
                        Category.objects.get_or_create(
                            name="Другое",
                            parent=child,
                            defaults={
                                'is_active': True,
                                'slug': f'other-{child.slug}' if child.slug else f'other-{child.id}'
                            }
                        )
    
    def has_change_permission(self, request, obj=None):
        """Проверка права на изменение продукта"""
        # Сначала проверяем базовые права
        if not super().has_change_permission(request, obj):
            return False
        
        # Суперпользователи могут менять любые продукты
        if request.user.is_superuser:
            return True
        
        # Для конкретного объекта проверяем принадлежность компании
        if obj:
            user_company = self.get_user_company(request)
            if user_company:
                return obj.company == user_company
            return False
        
        # Для общего доступа (список) разрешаем, если есть компания
        user_company = self.get_user_company(request)
        return user_company is not None
    
    def has_delete_permission(self, request, obj=None):
        """Проверка права на удаление продукта"""
        # Сначала проверяем базовые права
        if not super().has_delete_permission(request, obj):
            return False
        
        # Суперпользователи могут удалять любые продукты
        if request.user.is_superuser:
            return True
        
        # Для конкретного объекта проверяем принадлежность компании
        if obj:
            user_company = self.get_user_company(request)
            if user_company:
                return obj.company == user_company
            return False
        
        # Для общего доступа (список) разрешаем, если есть компания
        user_company = self.get_user_company(request)
        return user_company is not None
    
    def get_price_display(self, obj):
        if obj.price:
            return f"{obj.price} {obj.currency}"
        return "Договорная"
    get_price_display.short_description = "Цена"
    
    def get_price_conversions(self, obj):
        """Display price conversions in other currencies"""
        if not obj.price:
            return "-"
        
        conversions = []
        for currency in ['KZT', 'RUB', 'USD']:
            if currency != obj.currency:
                converted_price = obj.get_price_in(currency)
                if converted_price:
                    conversions.append(f"{converted_price} {currency}")
        
        if conversions:
            conversion_text = " | ".join(conversions)
            return format_html(
                '<span style="color: #666; font-size: 0.9em;">{}</span>', 
                conversion_text
            )
        return "-"
    get_price_conversions.short_description = "Конвертация валют"

    def test_xlsx_export_in_memory(self):
        """
        Тестовый метод для проверки работы XLSX экспорта в памяти
        Создает пример данных и экспортирует их в XLSX формат
        """
        from tablib import Dataset
        import io

        print("🧪 Начинаем тест XLSX экспорта в памяти...")

        # Создаем тестовый dataset с примером данных продуктов
        resource = ProductResource()

        # Получаем несколько продуктов для экспорта (максимум 3 для теста)
        queryset = Product.objects.all()[:3]

        if not queryset.exists():
            print("❌ Нет продуктов для тестирования экспорта")
            return False

        # Экспортируем данные в dataset
        dataset = resource.export(queryset)

        # Проверяем, что dataset содержит данные
        if not dataset.dict:
            print("❌ Dataset пустой")
            return False

        print(f"📊 Dataset содержит {len(dataset.dict)} строк")
        print(f"📋 Колонки: {dataset.headers}")

        # Конвертируем dataset в XLSX формат в памяти
        try:
            xlsx_data = dataset.export('xlsx')
            print(f"✅ XLSX данные созданы успешно, размер: {len(xlsx_data)} байт")

            # Дополнительная проверка - читаем данные обратно
            import openpyxl
            xlsx_file = io.BytesIO(xlsx_data)
            workbook = openpyxl.load_workbook(xlsx_file)
            worksheet = workbook.active

            print(f"📖 XLSX файл прочитан, активный лист: {worksheet.title}")
            print(f"📏 Размер листа: {worksheet.max_row} строк x {worksheet.max_column} колонок")

            # Показываем первую строку заголовков
            if worksheet.max_row > 0:
                headers = []
                for cell in worksheet[1]:
                    headers.append(str(cell.value) if cell.value else "")
                print(f"🏷️  Заголовки: {headers}")

            workbook.close()
            xlsx_file.close()

            print("✅ Тест XLSX экспорта в памяти прошел успешно!")
            return True

        except Exception as e:
            print(f"❌ Ошибка при создании XLSX: {str(e)}")
            return False

    def generate_sample_excel(self, request, queryset=None):
        """
        Генерация образца Excel файла для импорта продуктов
        """
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        # Создаём новый workbook для образца
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Образец продуктов"

        # Определяем заголовки колонок для импорта (согласно ProductResource)
        headers = ['ID', 'Компания', 'Название', 'Категория', 'Описание', 'Цена', 'Остаток', 'Активен']

        # Стилизация заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Добавляем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Добавляем примеры данных
        sample_data = [
            ['', 'ТОО "Электроника Плюс"', 'iPhone 15 Pro', 'Телефоны', 'Новейший смартфон Apple с улучшенной камерой', 450000, 10, 'да'],
            ['', 'ТОО "Электроника Плюс"', 'Samsung Galaxy S24', 'Телефоны', 'Флагманский Android-смартфон', 380000, 5, 'да'],
            ['', 'ТОО "КомпТех"', 'Ноутбук ASUS ROG', 'Компьютеры', 'Игровой ноутбук для профессионалов', 650000, 0, 'нет'],
            ['', 'ТОО "КомпТех"', 'Консультация по IT', 'Услуги', 'Консультация специалиста по IT-вопросам', 15000, '', 'да'],
        ]

        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Автоматическое изменение ширины колонок
        for col_num in range(1, len(headers) + 1):
            column = get_column_letter(col_num)
            max_length = 0
            for cell in worksheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Устанавливаем ширину с небольшим отступом
            adjusted_width = max(max_length + 2, 15)  # Минимальная ширина 15 символов
            worksheet.column_dimensions[column].width = adjusted_width

        # Добавляем инструкции на отдельный лист
        instructions_sheet = workbook.create_sheet("Инструкции")
        instructions = [
            "ИНСТРУКЦИЯ ПО ИМПОРТУ ПРОДУКТОВ",
            "",
            "Заполните таблицу на листе 'Образец продуктов' следующим образом:",
            "",
            "1. ID - оставьте пустым для новых продуктов или укажите существующий ID для обновления",
            "2. Компания - ОБЯЗАТЕЛЬНОЕ поле, точное название компании",
            "3. Название - ОБЯЗАТЕЛЬНОЕ поле, название продукта или услуги",
            "4. Категория - название категории (если не существует, будет создана автоматически)",
            "5. Описание - подробное описание продукта или услуги",
            "6. Цена - цена в тенге (можно оставить пустым для договорной цены)",
            "7. Остаток - количество на складе (число больше 0 = 'в наличии', 0 или меньше = 'нет в наличии')",
            "8. Активен - 'да' для публикации, 'нет' для скрытия (по умолчанию 'да')",
            "",
            "ПРИМЕЧАНИЯ:",
            "• Если компания не существует, она будет создана автоматически",
            "• Если категория не существует, она будет создана автоматически",
            "• Для услуг оставьте поле 'Остаток' пустым",
            "• Используйте только 'да'/'нет' в поле 'Активен'",
            "• Цена указывается в казахстанских тенге (KZT)",
            "",
            "ВАЖНО:",
            "• Сохраните файл в формате .xlsx (Excel)",
            "• Не изменяйте названия колонок в первой строке",
            "• Поля 'Компания' и 'Название' обязательны для заполнения",
            "• Убедитесь, что все данные заполнены корректно перед импортом",
        ]

        for row_num, instruction in enumerate(instructions, 1):
            cell = instructions_sheet.cell(row=row_num, column=1, value=instruction)
            if row_num == 1:  # Заголовок
                cell.font = Font(bold=True, size=14)
            elif instruction == "":  # Пустые строки для разделения
                continue
            elif instruction.startswith("ПРИМЕЧАНИЯ:") or instruction.startswith("ВАЖНО:"):
                cell.font = Font(bold=True, size=12)

        # Устанавливаем ширину колонки для инструкций
        instructions_sheet.column_dimensions['A'].width = 80

        # Создаём HTTP ответ с Excel файлом
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Генерируем имя файла
        filename = 'sample_products_import.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Сохраняем workbook в ответ
        workbook.save(response)

        return response

    generate_sample_excel.short_description = "Скачать образец Excel для импорта"

    def get_urls(self):
        """
        Добавляем URL для скачивания образца к стандартным URL админки
        """
        urls = super().get_urls()
        from django.urls import path

        custom_urls = [
            path(
                'sample-download/',
                self.admin_site.admin_view(self.generate_sample_excel),
                name='products_product_sample_download'
            ),
        ]
        return custom_urls + urls


@admin.register(ProductImage)
class ProductImageAdmin(admin.ModelAdmin):
    # Отображение полей в списке
    list_display = ["product", "image_thumbnail", "alt_text", "is_primary", "created_at"]
    list_filter = ["is_primary", "created_at"]
    search_fields = ["product__title", "alt_text"]

    # Поля для редактирования
    fields = ["product", "image", "alt_text", "is_primary"]

    def image_thumbnail(self, obj):
        """Отображение миниатюры изображения в списке"""
        if obj.image:
            return format_html(
                '<img src="{}" style="width: 50px; height: 50px; object-fit: cover; border-radius: 4px;" />',
                obj.image.url
            )
        return "Нет изображения"
    image_thumbnail.short_description = "Изображение"


