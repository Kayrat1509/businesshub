from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import generics, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.response import Response

# Импорты для админки
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponseRedirect
from django.urls import reverse
from openpyxl import load_workbook
import logging

from app.common.permissions import IsAdminOrReadOnly

from .models import Category
from .serializers import CategorySerializer, CategoryTreeSerializer
from .forms import CategoryImportForm

logger = logging.getLogger(__name__)


class CategoryListCreateView(generics.ListCreateAPIView):
    queryset = Category.objects.filter(is_active=True)
    serializer_class = CategorySerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ["parent", "is_active"]
    search_fields = ["name"]
    ordering_fields = ["name", "created_at"]
    ordering = ["name"]


class CategoryRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAdminOrReadOnly]
    lookup_field = "slug"


@api_view(["GET"])
@permission_classes([permissions.AllowAny])
def category_tree(request):
    root_categories = Category.objects.filter(parent=None, is_active=True)
    serializer = CategoryTreeSerializer(root_categories, many=True)
    return Response(serializer.data)


@staff_member_required
def import_categories_from_excel(request):
    """
    Представление для импорта категорий из Excel файла
    """
    if request.method == 'POST':
        form = CategoryImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = form.cleaned_data['excel_file']
                result = process_excel_import(excel_file)
                
                messages.success(
                    request, 
                    f'Импортировано {result["created"]} новых категорий, '
                    f'обновлено {result["updated"]} существующих категорий.'
                )
                
                # Перенаправляем обратно на страницу списка категорий
                return HttpResponseRedirect(reverse('admin:categories_category_changelist'))
                
            except Exception as e:
                logger.error(f"Ошибка при импорте категорий: {str(e)}")
                messages.error(request, f'Ошибка при импорте файла: {str(e)}')
    else:
        form = CategoryImportForm()
    
    context = {
        'form': form,
        'title': 'Импорт категорий из Excel',
        'site_header': 'Администрирование B2B платформы',
    }
    
    return render(request, 'admin/categories/import_form.html', context)


def process_excel_import(excel_file):
    """
    Обработка Excel файла и импорт категорий
    """
    # Загружаем Excel файл через openpyxl
    workbook = load_workbook(excel_file, read_only=True)
    worksheet = workbook.active
    
    created_count = 0
    updated_count = 0
    errors = []
    
    # Получаем заголовки из первой строки
    headers = []
    for cell in worksheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    # Создаем словарь заголовков в нижнем регистре для поиска
    headers_lower = {header.lower(): idx for idx, header in enumerate(headers)}
    
    # Проверяем наличие обязательных колонок (нечувствительно к регистру)
    required_columns = ['name', 'slug']  # В нижнем регистре
    for col in required_columns:
        if col not in headers_lower:
            raise ValueError(f"Отсутствует обязательная колонка: {col.upper()}")
    
    # Получаем индексы колонок (нечувствительно к регистру)
    name_idx = headers_lower.get('name')
    slug_idx = headers_lower.get('slug')
    parent_idx = headers_lower.get('parent', None)
    
    # Обрабатываем строки данных (начиная со второй строки)
    for row_num, row in enumerate(worksheet.iter_rows(min_row=2), 2):
        try:
            # Извлекаем данные из ячеек
            name = str(row[name_idx].value).strip() if row[name_idx].value else ""
            slug = str(row[slug_idx].value).strip() if row[slug_idx].value else ""
            parent_name = str(row[parent_idx].value).strip() if parent_idx is not None and row[parent_idx].value else ""
            
            # Пропускаем пустые строки
            if not name or not slug:
                continue
                
            # Ищем родительскую категорию, если указана
            parent_category = None
            if parent_name:
                try:
                    parent_category = Category.objects.get(name=parent_name)
                except Category.DoesNotExist:
                    # Если родительская категория не найдена, создаём её
                    parent_category = Category.objects.create(
                        name=parent_name,
                        slug=parent_name.lower().replace(' ', '-'),
                        is_active=True
                    )
                    created_count += 1
                    logger.info(f"Создана родительская категория: {parent_name}")
            
            # Проверяем существование категории по slug
            category, created = Category.objects.get_or_create(
                slug=slug,
                defaults={
                    'name': name,
                    'parent': parent_category,
                    'is_active': True
                }
            )
            
            if created:
                created_count += 1
                logger.info(f"Создана новая категория: {name}")
            else:
                # Обновляем существующую категорию
                category.name = name
                category.parent = parent_category
                category.save()
                updated_count += 1
                logger.info(f"Обновлена категория: {name}")
                
        except Exception as e:
            error_msg = f"Ошибка в строке {row_num}: {str(e)}"
            errors.append(error_msg)
            logger.error(error_msg)
            continue
    
    workbook.close()
    
    if errors:
        error_summary = "\n".join(errors[:10])  # Показываем только первые 10 ошибок
        if len(errors) > 10:
            error_summary += f"\n... и ещё {len(errors) - 10} ошибок"
        raise ValueError(f"Обнаружены ошибки при импорте:\n{error_summary}")
    
    return {
        'created': created_count,
        'updated': updated_count
    }
