from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import datetime

from .models import Category


class ModerationStatusFilter(admin.SimpleListFilter):
    title = 'статус модерации'
    parameter_name = 'moderation_status'

    def lookups(self, request, model_admin):
        return (
            ('pending', '⏳ Ожидают модерации'),
            ('approved', '✓ Одобренные'),
        )

    def queryset(self, request, queryset):
        if self.value() == 'pending':
            return queryset.filter(is_active=False)
        if self.value() == 'approved':
            return queryset.filter(is_active=True)
        return queryset


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ["display_name", "parent", "slug", "status_display", "is_active", "products_count", "companies_count", "created_at"]
    list_filter = [
        ModerationStatusFilter,
        "parent",
        ("created_at", admin.DateFieldListFilter),
    ]
    search_fields = ["name", "slug"]
    prepopulated_fields = {"slug": ("name",)}
    list_editable = ["is_active"]
    ordering = ["is_active", "parent__name", "name"]  # Неактивные (на модерации) сверху
    list_per_page = 50
    actions = ['export_to_excel', 'approve_categories', 'reject_categories']

    # Указываем наш кастомный шаблон для списка
    change_list_template = 'admin/categories/category/change_list.html'

    fieldsets = (
        ("Основная информация", {
            "fields": ("name", "slug", "parent", "is_active")
        }),
        ("Служебная информация", {
            "fields": ("created_at", "updated_at"),
            "classes": ("collapse",)
        }),
    )
    readonly_fields = ["created_at", "updated_at"]

    def display_name(self, obj):
        if obj.parent:
            return format_html("&nbsp;&nbsp;&nbsp;&nbsp;└─ {}", obj.name)
        return format_html("<strong>{}</strong>", obj.name)
    display_name.short_description = "Название"
    display_name.admin_order_field = "name"

    def status_display(self, obj):
        if obj.is_active:
            return format_html(
                '<span style="color: green; font-weight: bold;">✓ Одобрено</span>'
            )
        else:
            return format_html(
                '<span style="color: orange; font-weight: bold;">⏳ На модерации</span>'
            )
    status_display.short_description = "Статус"
    status_display.admin_order_field = "is_active"

    def products_count(self, obj):
        count = obj.products.count()
        if count > 0:
            return format_html('<span style="color: green; font-weight: bold;">{}</span>', count)
        return count
    products_count.short_description = "Продукты"

    def companies_count(self, obj):
        from app.companies.models import Company
        count = Company.objects.filter(categories=obj).count()
        if count > 0:
            return format_html('<span style="color: blue; font-weight: bold;">{}</span>', count)
        return count
    companies_count.short_description = "Компании"

    def export_to_excel(self, request, queryset):
        """
        Экспорт выбранных категорий в Excel файл
        """
        # Создаём новый workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Категории"
        
        # Определяем заголовки колонок
        headers = ['ID', 'Название', 'Slug', 'Родитель']
        
        # Стилизация заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавляем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Добавляем данные категорий
        for row_num, category in enumerate(queryset, 2):
            worksheet.cell(row=row_num, column=1, value=category.id)
            worksheet.cell(row=row_num, column=2, value=category.name)
            worksheet.cell(row=row_num, column=3, value=category.slug)
            # Для родительской категории показываем name, если есть, иначе пусто
            parent_name = category.parent.name if category.parent else ""
            worksheet.cell(row=row_num, column=4, value=parent_name)
        
        # Автоматическое изменение ширины колонок
        for col_num in range(1, len(headers) + 1):
            column = get_column_letter(col_num)
            max_length = 0
            for cell in worksheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Устанавливаем ширину с небольшим отступом
            adjusted_width = min(max_length + 2, 50)  # Максимальная ширина 50 символов
            worksheet.column_dimensions[column].width = adjusted_width
        
        # Создаём HTTP ответ с Excel файлом
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Генерируем имя файла с текущей датой
        current_date = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'categories_{current_date}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Сохраняем workbook в ответ
        workbook.save(response)
        
        # Сообщение об успешном экспорте (если middleware доступен)
        try:
            self.message_user(request, f'Экспортировано {queryset.count()} категорий в файл {filename}')
        except:
            # Игнорируем ошибку если middleware недоступен
            pass
        
        return response
    
    export_to_excel.short_description = "Экспорт в Excel"

    def approve_categories(self, request, queryset):
        """
        Одобрить выбранные категории (активировать)
        """
        updated = queryset.filter(is_active=False).update(is_active=True)
        if updated:
            self.message_user(
                request,
                f'Одобрено {updated} категорий. Они теперь доступны в каталоге.'
            )
        else:
            self.message_user(
                request,
                'Нет категорий для одобрения (все выбранные уже активны).'
            )
    approve_categories.short_description = "✓ Одобрить выбранные категории"

    def reject_categories(self, request, queryset):
        """
        Отклонить выбранные категории (удалить)
        """
        pending_categories = queryset.filter(is_active=False)
        count = pending_categories.count()

        if count > 0:
            # Проверяем, не связаны ли категории с продуктами
            categories_with_products = []
            for category in pending_categories:
                if category.products.exists():
                    categories_with_products.append(category.name)

            if categories_with_products:
                self.message_user(
                    request,
                    f'Нельзя удалить категории, так как они используются в продуктах: {", ".join(categories_with_products)}',
                    level='ERROR'
                )
                return

            # Удаляем категории без связанных продуктов
            pending_categories.delete()
            self.message_user(
                request,
                f'Отклонено и удалено {count} категорий.'
            )
        else:
            self.message_user(
                request,
                'Нет категорий для отклонения (можно отклонять только неактивные категории).'
            )
    reject_categories.short_description = "✗ Отклонить выбранные категории"

    def generate_sample_excel(self, request, queryset=None):
        """
        Генерация образца Excel файла для импорта категорий
        """
        # Создаём новый workbook для образца
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Образец категорий"

        # Определяем заголовки колонок для импорта
        headers = ['ID', 'Название', 'Slug', 'Родитель']

        # Стилизация заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Добавляем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Добавляем примеры данных
        sample_data = [
            ['', 'Электроника', 'elektronika', ''],  # Пример родительской категории
            ['', 'Телефоны', 'telefony', 'Электроника'],  # Пример дочерней категории
            ['', 'Компьютеры', 'kompyutery', 'Электроника'],  # Еще один пример
            ['', 'Одежда', 'odezhda', ''],  # Еще одна родительская категория
        ]

        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Автоматическое изменение ширины колонок
        for col_num in range(1, len(headers) + 1):
            column = get_column_letter(col_num)
            max_length = 0
            for cell in worksheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Устанавливаем ширину с небольшим отступом
            adjusted_width = max(max_length + 2, 15)  # Минимальная ширина 15 символов
            worksheet.column_dimensions[column].width = adjusted_width

        # Добавляем инструкции на отдельный лист
        instructions_sheet = workbook.create_sheet("Инструкции")
        instructions = [
            "ИНСТРУКЦИЯ ПО ИМПОРТУ КАТЕГОРИЙ",
            "",
            "Заполните таблицу на листе 'Образец категорий' следующим образом:",
            "",
            "1. ID - оставьте пустым для новых категорий или укажите существующий ID для обновления",
            "2. Название - обязательное поле, название категории на русском языке",
            "3. Slug - транслитерация названия латинскими буквами (можно оставить пустым)",
            "4. Родитель - название родительской категории (если это подкategория)",
            "",
            "ПРИМЕЧАНИЯ:",
            "• Для создания родительской категории оставьте поле 'Родитель' пустым",
            "• Для создания подкategorии укажите название родительской категории",
            "• Если родительская категория не существует, она будет создана автоматически",
            "• Slug генерируется автоматически из названия, если не указан",
            "",
            "ВАЖНО:",
            "• Сохраните файл в формате .xlsx (Excel)",
            "• Не изменяйте названия колонок в первой строке",
            "• Убедитесь, что все данные заполнены корректно перед импортом",
        ]

        for row_num, instruction in enumerate(instructions, 1):
            cell = instructions_sheet.cell(row=row_num, column=1, value=instruction)
            if row_num == 1:  # Заголовок
                cell.font = Font(bold=True, size=14)
            elif instruction == "":  # Пустые строки для разделения
                continue
            elif instruction.startswith("ПРИМЕЧАНИЯ:") or instruction.startswith("ВАЖНО:"):
                cell.font = Font(bold=True, size=12)

        # Устанавливаем ширину колонки для инструкций
        instructions_sheet.column_dimensions['A'].width = 80

        # Создаём HTTP ответ с Excel файлом
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Генерируем имя файла
        filename = 'sample_categories_import.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Сохраняем workbook в ответ
        workbook.save(response)

        return response

    generate_sample_excel.short_description = "Скачать образец Excel для импорта"

    def get_urls(self):
        """
        Добавляем URL для скачивания образца к стандартным URL админки
        """
        urls = super().get_urls()
        from django.urls import path

        custom_urls = [
            path(
                'sample-download/',
                self.admin_site.admin_view(self.generate_sample_excel),
                name='categories_category_sample_download'
            ),
        ]
        return custom_urls + urls

