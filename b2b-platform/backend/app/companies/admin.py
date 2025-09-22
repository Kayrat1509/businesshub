from django.contrib import admin
from django import forms
from django.forms import Textarea, TextInput
from django.urls import reverse
from django.utils.safestring import mark_safe
from import_export.admin import ImportExportModelAdmin
import json

from .models import Branch, Company, Employee
from .resources import CompanyResource



class CompanyAdminForm(forms.ModelForm):
    # Контакты
    phone_numbers = forms.CharField(
        label="Номера телефонов", 
        max_length=500, 
        required=False,
        help_text="Введите номера телефонов через запятую, например: +7 (999) 123-45-67, +7 (999) 765-43-21",
        widget=TextInput(attrs={'size': '60'})
    )
    emails = forms.CharField(
        label="Email адреса", 
        max_length=500, 
        required=False,
        help_text="Введите email адреса через запятую, например: info@company.com, sales@company.com",
        widget=TextInput(attrs={'size': '60'})
    )
    website = forms.URLField(
        label="Веб-сайт", 
        required=False,
        help_text="Введите адрес веб-сайта, например: https://company.com",
        widget=TextInput(attrs={'size': '60'})
    )
    
    # Социальные сети
    facebook = forms.URLField(
        label="Facebook", 
        required=False,
        help_text="Ссылка на Facebook страницу",
        widget=TextInput(attrs={'size': '60'})
    )
    instagram = forms.URLField(
        label="Instagram", 
        required=False,
        help_text="Ссылка на Instagram профиль",
        widget=TextInput(attrs={'size': '60'})
    )
    telegram = forms.URLField(
        label="Telegram", 
        required=False,
        help_text="Ссылка на Telegram канал или бот",
        widget=TextInput(attrs={'size': '60'})
    )
    whatsapp = forms.URLField(
        label="WhatsApp", 
        required=False,
        help_text="Ссылка на WhatsApp Business",
        widget=TextInput(attrs={'size': '60'})
    )
    twitter = forms.URLField(
        label="Twitter", 
        required=False,
        help_text="Ссылка на Twitter профиль",
        widget=TextInput(attrs={'size': '60'})
    )
    linkedin = forms.URLField(
        label="LinkedIn", 
        required=False,
        help_text="Ссылка на LinkedIn компании",
        widget=TextInput(attrs={'size': '60'})
    )
    
    # Юридическая информация
    inn = forms.CharField(
        label="ИНН", 
        max_length=12, 
        required=False,
        help_text="Индивидуальный налоговый номер (10 или 12 цифр)",
        widget=TextInput(attrs={'size': '20'})
    )
    kpp = forms.CharField(
        label="КПП", 
        max_length=9, 
        required=False,
        help_text="Код причины постановки на учет (9 цифр)",
        widget=TextInput(attrs={'size': '20'})
    )
    legal_name = forms.CharField(
        label="Полное наименование организации", 
        max_length=300, 
        required=False,
        help_text="Официальное полное наименование организации",
        widget=TextInput(attrs={'size': '80'})
    )
    
    # Способы оплаты
    accepts_cash = forms.BooleanField(label="Принимаем наличные", required=False)
    accepts_cards = forms.BooleanField(label="Принимаем банковские карты", required=False)
    accepts_transfers = forms.BooleanField(label="Принимаем банковские переводы", required=False)
    accepts_crypto = forms.BooleanField(label="Принимаем криптовалюту", required=False)
    
    # График работы
    work_hours = forms.CharField(
        label="График работы",
        max_length=200,
        required=False,
        help_text="Например: Пн-Пт 9:00-18:00, Сб 10:00-16:00, Вс - выходной",
        widget=TextInput(attrs={'size': '60'})
    )

    # Координаты
    coordinates = forms.CharField(
        label="Широта и Долгота",
        max_length=100,
        required=False,
        help_text="Введите координаты в формате: широта, долгота (например: 47.05505112306978, 51.82543208082456)",
        widget=TextInput(attrs={'size': '60', 'placeholder': '47.05505112306978, 51.82543208082456'})
    )

    class Meta:
        model = Company
        exclude = ['contacts', 'legal_info', 'payment_methods', 'work_schedule', 'latitude', 'longitude']
        # Включаем новые поля phones и supplier_type в форму

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        if self.instance and self.instance.pk:
            # Редактирование существующей компании - заполнение полей из JSON данных
            contacts = self.instance.contacts or {}
            legal_info = self.instance.legal_info or {}
            payment_methods = self.instance.payment_methods or []
            work_schedule = self.instance.work_schedule or {}

            # Собираем все телефоны из разных источников
            all_phones = []

            # Добавляем одинарный телефон из contacts.phone
            if contacts.get('phone'):
                all_phones.append(contacts['phone'])

            # Добавляем массив телефонов из contacts.phones
            if contacts.get('phones') and isinstance(contacts['phones'], list):
                for phone in contacts['phones']:
                    if phone and phone not in all_phones:  # Избегаем дублирования
                        all_phones.append(phone)

            # Собираем все email адреса из разных источников
            all_emails = []

            # Добавляем одинарный email из contacts.email
            if contacts.get('email'):
                all_emails.append(contacts['email'])

            # Добавляем массив emails из contacts.emails
            if contacts.get('emails') and isinstance(contacts['emails'], list):
                for email in contacts['emails']:
                    if email and email not in all_emails:  # Избегаем дублирования
                        all_emails.append(email)

            self.fields['phone_numbers'].initial = ', '.join(all_phones)
            self.fields['emails'].initial = ', '.join(all_emails)
            self.fields['website'].initial = contacts.get('website', '')
            
            # Социальные сети
            social = contacts.get('social', {})
            self.fields['facebook'].initial = social.get('facebook', '')
            self.fields['instagram'].initial = social.get('instagram', '')
            self.fields['telegram'].initial = social.get('telegram', '')
            self.fields['whatsapp'].initial = social.get('whatsapp', '')
            self.fields['twitter'].initial = social.get('twitter', '')
            self.fields['linkedin'].initial = social.get('linkedin', '')
            
            self.fields['inn'].initial = legal_info.get('inn', '')
            self.fields['kpp'].initial = legal_info.get('kpp', '')
            self.fields['legal_name'].initial = legal_info.get('legal_name', '')
            
            self.fields['accepts_cash'].initial = 'CASH' in payment_methods
            self.fields['accepts_cards'].initial = 'CARD' in payment_methods
            self.fields['accepts_transfers'].initial = 'TRANSFER' in payment_methods
            self.fields['accepts_crypto'].initial = 'CRYPTO' in payment_methods
            
            self.fields['work_hours'].initial = work_schedule.get('description', '')

            # Координаты
            if self.instance.latitude and self.instance.longitude:
                self.fields['coordinates'].initial = f"{self.instance.latitude}, {self.instance.longitude}"
        else:
            # Создание новой компании - устанавливаем дефолтное значение для телефона
            self.fields['phone_numbers'].initial = '+7 (777) 123-45-67'

    def save(self, commit=True):
        instance = super().save(commit=False)
        
        # Set default status to APPROVED for admin created companies
        if not instance.pk and not hasattr(instance, '_state'):
            instance.status = 'APPROVED'
        
        # Сохранение контактов в JSON
        phones = [phone.strip() for phone in self.cleaned_data['phone_numbers'].split(',') if phone.strip()]
        emails = [email.strip() for email in self.cleaned_data['emails'].split(',') if email.strip()]
        website = self.cleaned_data['website']
        
        # Социальные сети
        social_media = {}
        for platform in ['facebook', 'instagram', 'telegram', 'whatsapp', 'twitter', 'linkedin']:
            if self.cleaned_data[platform]:
                social_media[platform] = self.cleaned_data[platform]
        
        # Сохраняем контакты в обеих форматах для совместимости
        contacts_data = {
            'phones': phones,  # Массив телефонов для админки
            'emails': emails,  # Массив emails для админки
            'website': website,
            'social': social_media
        }

        # Добавляем одинарные поля для совместимости с личным кабинетом
        if phones:
            contacts_data['phone'] = phones[0]  # Первый телефон как основной
        if emails:
            contacts_data['email'] = emails[0]  # Первый email как основной

        instance.contacts = contacts_data
        
        # Сохранение юридической информации в JSON
        instance.legal_info = {
            'inn': self.cleaned_data['inn'],
            'kpp': self.cleaned_data['kpp'],
            'legal_name': self.cleaned_data['legal_name']
        }
        
        # Сохранение способов оплаты в JSON
        payment_methods = []
        if self.cleaned_data['accepts_cash']:
            payment_methods.append('CASH')
        if self.cleaned_data['accepts_cards']:
            payment_methods.append('CARD')
        if self.cleaned_data['accepts_transfers']:
            payment_methods.append('TRANSFER')
        if self.cleaned_data['accepts_crypto']:
            payment_methods.append('CRYPTO')
        
        instance.payment_methods = payment_methods
        
        # Сохранение графика работы в JSON
        instance.work_schedule = {
            'description': self.cleaned_data['work_hours']
        }

        # Парсинг и сохранение координат
        coordinates_str = self.cleaned_data.get('coordinates', '').strip()
        if coordinates_str:
            try:
                coords = [coord.strip() for coord in coordinates_str.split(',')]
                if len(coords) == 2:
                    lat = float(coords[0])
                    lng = float(coords[1])
                    # Проверяем что координаты в допустимых диапазонах
                    if -90 <= lat <= 90 and -180 <= lng <= 180:
                        instance.latitude = lat
                        instance.longitude = lng
                    else:
                        # Если координаты некорректные, очищаем их
                        instance.latitude = None
                        instance.longitude = None
                else:
                    # Если формат некорректный, очищаем координаты
                    instance.latitude = None
                    instance.longitude = None
            except (ValueError, TypeError):
                # Если парсинг не удался, очищаем координаты
                instance.latitude = None
                instance.longitude = None
        else:
            # Если поле пустое, очищаем координаты
            instance.latitude = None
            instance.longitude = None

        if commit:
            instance.save()
        return instance


@admin.register(Company)
class CompanyAdmin(ImportExportModelAdmin):
    # Подключаем ресурс для импорта/экспорта
    resource_class = CompanyResource
    form = CompanyAdminForm

    # Ограничиваем форматы импорта/экспорта только Excel (.xlsx)
    from import_export.formats.base_formats import XLSX
    formats = [XLSX]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Принудительно устанавливаем наш шаблон после инициализации ImportExportModelAdmin
        self.change_list_template = 'admin/companies/company/change_list.html'
    
    list_display = ["name", "owner", "supplier_type", "city", "status", "rating", "created_at"]
    list_filter = ["status", "supplier_type", "city", "categories", "created_at"]
    search_fields = ["name", "description", "phones", "owner__email"]
    list_editable = ["status", "supplier_type"]
    filter_horizontal = ["categories"]
    readonly_fields = ["rating", "created_at", "updated_at"]

    fieldsets = (
        (
            "Основная информация",
            {"fields": ("name", "owner", "supplier_type", "logo", "description", "categories")},
        ),
        (
            "Контактная информация",
            {"fields": ("phone_numbers", "emails", "website", "phones")},
        ),
        (
            "Социальные сети",
            {"fields": ("facebook", "instagram", "telegram", "whatsapp", "twitter", "linkedin"), "classes": ["collapse"]},
        ),
        (
            "Местоположение",
            {"fields": ("city", "address", "coordinates")},
        ),
        (
            "Юридическая информация",
            {"fields": ("legal_name", "inn", "kpp")},
        ),
        (
            "Способы оплаты",
            {"fields": ("accepts_cash", "accepts_cards", "accepts_transfers", "accepts_crypto")},
        ),
        (
            "График работы и детали",
            {"fields": ("work_hours", "staff_count", "branches_count")},
        ),
        ("Статус и рейтинг", {"fields": ("status", "rating")}),
        (
            "Временные метки",
            {"fields": ("created_at", "updated_at"), "classes": ["collapse"]},
        ),
    )

    # Удалили changelist_view так как теперь используем change_list_template

    def generate_sample_excel(self, request, queryset=None):
        """
        Генерация образца Excel файла для импорта компаний
        """
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        # Создаём новый workbook для образца
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Образец компаний"

        # Определяем заголовки колонок для импорта (согласно CompanyResource)
        headers = ['ID', 'Название', 'Номера телефонов', 'Описание', 'Город', 'Адрес',
                  'Тип поставщика', 'Контакты', 'Юр. информация', 'Способы оплаты',
                  'График работы', 'Статус', 'Владелец', 'Дата создания']

        # Стилизация заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Добавляем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Добавляем примеры данных
        sample_data = [
            ['', 'ТОО "Электроника Плюс"', '+7 (777) 123-45-67, +7 (777) 987-65-43',
             'Продажа электроники и мобильных устройств', 'Алматы', 'пр. Абая, 150',
             'Дилер', '{"phones": ["+7 (777) 123-45-67"], "emails": ["info@electroplus.kz"], "website": "https://electroplus.kz"}',
             '{"inn": "123456789012", "legal_name": "ТОО Электроника Плюс"}', '["CASH", "CARD", "TRANSFER"]',
             '{"description": "Пн-Пт 9:00-18:00, Сб 10:00-16:00"}', 'Одобрено', 'admin@site.com', ''],

            ['', 'ТОО "КомпТех"', '+7 (707) 555-12-34',
             'IT-консалтинг и разработка программного обеспечения', 'Нур-Султан', 'ул. Сауран, 12/1',
             'Производитель', '{"phones": ["+7 (707) 555-12-34"], "emails": ["contact@komptech.kz"]}',
             '{"inn": "987654321098", "legal_name": "ТОО КомпТех"}', '["CARD", "TRANSFER"]',
             '{"description": "Пн-Пт 9:00-18:00"}', 'Одобрено', 'admin@site.com', ''],
        ]

        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Автоматическое изменение ширины колонок
        for col_num in range(1, len(headers) + 1):
            column = get_column_letter(col_num)
            max_length = 0
            for cell in worksheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Устанавливаем ширину с небольшим отступом
            adjusted_width = max(max_length + 2, 20)  # Минимальная ширина 20 символов для компаний
            worksheet.column_dimensions[column].width = min(adjusted_width, 50)  # Максимум 50

        # Добавляем инструкции на отдельный лист
        instructions_sheet = workbook.create_sheet("Инструкции")
        instructions = [
            "ИНСТРУКЦИЯ ПО ИМПОРТУ КОМПАНИЙ",
            "",
            "Заполните таблицу на листе 'Образец компаний' следующим образом:",
            "",
            "1. ID - оставьте пустым для новых компаний или укажите существующий ID для обновления",
            "2. Название - ОБЯЗАТЕЛЬНОЕ поле, название компании",
            "3. Номера телефонов - основные телефоны компании через запятую",
            "4. Описание - краткое описание деятельности компании",
            "5. Город - город нахождения компании",
            "6. Адрес - полный адрес компании",
            "7. Тип поставщика - один из: Дилер, Производитель, Торговый представитель",
            "8. Контакты - JSON с дополнительными контактами (можно оставить пустым)",
            "9. Юр. информация - JSON с ИНН и юр. названием (можно оставить пустым)",
            "10. Способы оплаты - JSON массив способов оплаты (можно оставить пустым)",
            "11. График работы - JSON с описанием графика (можно оставить пустым)",
            "12. Статус - один из: Черновик, На модерации, Одобрено, Заблокировано",
            "13. Владелец - email владельца (если не указан, назначается admin)",
            "14. Дата создания - оставьте пустым (заполнится автоматически)",
            "",
            "ПРИМЕЧАНИЯ:",
            "• JSON поля можно оставить пустыми - они заполнятся значениями по умолчанию",
            "• Если владелец не найден, будет назначен суперпользователь",
            "• Новые компании автоматически получают статус 'Одобрено'",
            "• Поддерживаются как русские, так и английские названия статусов",
            "",
            "ВАЖНО:",
            "• Сохраните файл в формате .xlsx (Excel)",
            "• Не изменяйте названия колонок в первой строке",
            "• Поле 'Название' обязательно для заполнения",
            "• Убедитесь, что все данные заполнены корректно перед импортом",
        ]

        for row_num, instruction in enumerate(instructions, 1):
            cell = instructions_sheet.cell(row=row_num, column=1, value=instruction)
            if row_num == 1:  # Заголовок
                cell.font = Font(bold=True, size=14)
            elif instruction == "":  # Пустые строки для разделения
                continue
            elif instruction.startswith("ПРИМЕЧАНИЯ:") or instruction.startswith("ВАЖНО:"):
                cell.font = Font(bold=True, size=12)

        # Устанавливаем ширину колонки для инструкций
        instructions_sheet.column_dimensions['A'].width = 80

        # Создаём HTTP ответ с Excel файлом
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Генерируем имя файла
        filename = 'sample_companies_import.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Сохраняем workbook в ответ
        workbook.save(response)

        return response

    generate_sample_excel.short_description = "Скачать образец Excel для импорта"

    def get_urls(self):
        """
        Добавляем URL для скачивания образца к стандартным URL админки
        """
        urls = super().get_urls()
        from django.urls import path

        custom_urls = [
            path(
                'sample-download/',
                self.admin_site.admin_view(self.generate_sample_excel),
                name='companies_company_sample_download'
            ),
        ]
        return custom_urls + urls


@admin.register(Branch)
class BranchAdmin(admin.ModelAdmin):
    list_display = ["company", "address", "phone", "created_at"]
    list_filter = ["company", "created_at"]
    search_fields = ["company__name", "address", "phone"]


@admin.register(Employee)
class EmployeeAdmin(admin.ModelAdmin):
    list_display = ["full_name", "company", "position", "phone", "email"]
    list_filter = ["company", "position", "created_at"]
    search_fields = ["full_name", "company__name", "position", "email"]
